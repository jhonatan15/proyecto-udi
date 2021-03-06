from flask import Flask, render_template, request, flash, redirect, url_for, send_file
from openpyxl import load_workbook, Workbook
from datetime import datetime

app = Flask(__name__)

# Variables for fields of the form
fullname = ""
lastname = ""
ident = ""
birth = ""
status = ""
company = ""
position = ""
drugs = ""
disorder = ""
disease = ""

# Index route
@app.route('/')
def Index():
    return render_template('index.html', fullname = fullname, lastname = lastname, ident = ident, birth = birth, status = status, company = company, position = position, drugs = drugs, disorder = disorder)


# Global variables
now = datetime.now()
year = now.year
month = now.month
day = now.day

# Excel for records
wb_registro = load_workbook('respuestas/registro/registro.xlsx')
sheet_registro = wb_registro['Registros']
beginrow_registro = 2
finalrow_registro = 1000
listB_registro = [sheet_registro['A' + str(i)].value for i in range(beginrow_registro , finalrow_registro + 1)]

# Form
@app.route('/datos', methods=['POST'])
def datos():
    if request.method == 'POST':
        global fullname, lastname, ident, birth, status, company, position, drugs, disorder, disease

# Capture form information
        fullname = request.form['fullname']
        lastname = request.form['lastname']
        ident = request.form['ident']
        button = request.form['button']
        birth = request.form['birth']
        status = request.form['status']
        company = request.form['company']
        position = request.form['position']
        drugs = request.form['drugs']
        disorder = request.form['disorder']
        disease = request.form['disease']

# Confirm if the variables are full to continue to the second page
        if button == 'continuar':
            if fullname != "" and lastname != "" and ident != "" and birth != "" and status != "" and company != "" and position != "" and drugs != "" and disorder != "" and disease != "":
                number = 1

                # list of rows
                row_name = sheet_registro['B']
                for i in row_name:
                    if i.value is not None:
                        number += 1
                    elif i.value is None:
                        rowname = "B" + str(number)
                        rowlastname = "C" + str(number)
                        rowident = "D" + str(number)
                        rowbirth = "E" + str(number)
                        rowstatus = "F" + str(number)
                        rowcompany = "G" + str(number)
                        rowposition = "H" + str(number)
                        rowdrugs = "I" + str(number)
                        rowdisorder = "J" + str(number)
                        rowdisease = "K" + str(number)

                        sheet_registro[rowname] = fullname
                        sheet_registro[rowlastname] = lastname
                        sheet_registro[rowident] = ident
                        sheet_registro[rowbirth] = birth
                        sheet_registro[rowstatus] = status
                        sheet_registro[rowcompany] = company
                        sheet_registro[rowposition] = position
                        sheet_registro[rowdrugs] = drugs
                        sheet_registro[rowdisorder] = disorder
                        sheet_registro[rowdisease] = disease
                wb_registro.save('respuestas/registro/registro.xlsx')
                file_from = 'respuestas/registro/registro.xlsx'
                file_to = '/registro1/registro.xlsx'

                return render_template('consentimiento.html', fullname = fullname, lastname = lastname, ident = ident, year = year, month = month, day = day)

# Terms and conditions
@app.route('/consentimiento', methods=['POST'])
def consentimiento():
    if request.method == 'POST':
        global fullname, lastname
        button = request.form['button']
        if button == 'accept':
            return render_template('exam.html')

# Save answers of the first examn in excel
wb_first = load_workbook('primerexamen.xlsm')
listB2_first = ""
sheet_first = wb_first['Respuestas']
multiple_cells_first = sheet_first['B3':'B23']
beginrow_first = 3
finalrow_first = 23

@app.route('/exam', methods=['POST'])
def exam():
    if request.method == 'POST':

# Global variables
        global wb_first, sheet_first, listB2_first, fullname, lastname

# Get values ??????from a range of cells in excel
        listb_first = [sheet_first['B' + str(i)].value for i in range(beginrow_first , finalrow_first + 1)]
        sheet_first['F9'] = fullname
        sheet_first['G9'] = ident
        print(ident)
        try:
            question1 = request.form["question1"]
            question2 = request.form["question2"]
            question3 = request.form["question3"]
            question4 = request.form["question4"]
            question5 = request.form["question5"]
            question6 = request.form["question6"]
            question7 = request.form["question7"]
            question8 = request.form["question8"]
            question9 = request.form["question9"]
            question10 = request.form["question10"]
            question11 = request.form["question11"]
        except KeyError:
            pass
        button = request.form["button"]
        button2 = int(button)

# Function to for cycle
        def fill_excel():
            for i in listb_first:
                    ist = str(i)
                    if ist == number:
                        sheet2 = "C" + number
                        sheet_first[sheet2] = letter

# For cycle to separate numbers and letters from the first list and then write in their respective excel cell
        if button2 == 21:
            list_quest = [question1, question2, question3, question4, question5, question6, question7]
            list_quest2 = [question8, question9, question10]
            for i in list_quest:
                number = i[0]
                letter = i[2]
                fill_excel()
            for i in list_quest2:
                number = i[0:2]
                letter = i[3]
                fill_excel()

        if button2 == 22:
            print(listb_first)
            list_quest2 = [question1, question2, question3, question4, question5, question6, question7, question8, question9, question10, question11]
            for i in list_quest2:
                number = i[0:2]
                letter = i[3]
                fill_excel()

        wb_first.save('respuestas/encuesta1/1_'+ str(ident) +'.xlsx')
        file_from_2 = ('respuestas/encuesta1/1_'+ str(ident) +'.xlsx')
        file_to_2 = ('/respuestas1/1_'+ str(ident) +'.xlsx')
        if button2 == 21:
            return render_template('exam_21.html')
        elif button2 == 22:
            return render_template('exam_22.html')


# ---------------------------- # ---------------------------
# Open excel file for data entry
wb = load_workbook('sacarpuntajes.xlsm')
listB2 = ""
sheet = wb['Respuestas']
sheet_id = wb['id']
multiple_cells = sheet['B3':'B189']
beginrow = 3
finalrow = 189

# ----------------- Global Variables -------------------
fullname_2 = ""
ident_2 = ""

@app.route('/form_2')
def form_2():
    return render_template('datos_2.html', fullname_2 = fullname_2, ident_2 = ident_2)

@app.route('/datos_2', methods=['POST'])
def datos_2():
    if request.method == 'POST':
        global fullname_2
        global ident_2

# Capture form information
        fullname_2 = request.form['fullname_2']
        ident_2 = request.form['ident_2']
        button_2 = request.form['button']

# confirm if the variables are full to continue to the second page
        if button_2 == 'continuar2':
            if fullname_2 != "" and ident_2 != "":
                return render_template('instrucciones.html')

 # Instructions and examples
@app.route('/instrucciones', methods=['POST'])
def exam_1():
    if request.method == 'POST':
        button = request.form['button']
        if button == 'accept':
            return render_template('exam_1.html')

# Start exam

@app.route('/exam_1', methods=['POST'])
def exam_2():
    if request.method == 'POST':

# Global variables
        global wb, sheet, listB, wb2, fullname_2, ident_2

        sheet_id['B6'] = fullname_2
        sheet_id['C6'] = ident_2

# Get values ??????from a range of cells in excel
        listb = [sheet['B' + str(i)].value for i in range(beginrow , finalrow + 1)]
        try:
            question1 = request.form["question1"]
            question2 = request.form["question2"]
            question3 = request.form["question3"]
            question4 = request.form["question4"]
            question5 = request.form["question5"]
            question6 = request.form["question6"]
            question7 = request.form["question7"]
            question8 = request.form["question8"]
            question9 = request.form["question9"]
            question10 = request.form["question10"]
            question11 = request.form["question11"]
        except KeyError:
            pass
        button = request.form["button"]
        button2 = int(button)

# Function to for cycle
        def fill_excel_2():
            for i in listb:
                    ist = str(i)
                    if ist == number:
                        sheet2 = "C" + number
                        sheet[sheet2] = letter              
# For cycle to separate numbers and letters from the first list and then write in their respective excel cell
        if button2 == 2:
            list_quest = [question1, question2, question3, question4, question5, question6, question7]
            list_quest2 = [question8, question9, question10]
            for i in list_quest:
                number = i[0]
                letter = i[2]
                fill_excel_2()

            for i in list_quest2:
                number = i[0:2]
                letter = i[3]
                fill_excel_2()


        if button2 > 2 and button2  < 11:
            list_quest2 = [question1, question2, question3, question4, question5, question6, question7, question8, question9, question10]
            for i in list_quest2:
                number = i[0:2]
                letter = i[3]
                fill_excel_2()


        if button2 == 11:
            list_quest = [question1, question2, question3, question4, question5, question6, question7]
            list_quest2 = [question8, question9, question10]
            for i in list_quest:
                number = i[0:2]
                letter = i[3]
                fill_excel_2()

            for i in list_quest2:
                number = i[0:3]
                letter = i[4]
                fill_excel_2()


        if button2 > 11:
            if button2 == 20:
                list_quest = [question1, question2, question3, question4, question5, question6, question7]
                for i in list_quest:
                    number = i[0:3]
                    letter = i[4]
                    fill_excel_2()

            else:
                list_quest2 = [question1, question2, question3, question4, question5, question6, question7, question8, question9, question10]
                for i in list_quest2:
                    number = i[0:3]
                    letter = i[4]
                    fill_excel_2()





# Save excel file with the id of the person who filled out the form
        wb.save('respuestas/encuesta2/0_'+ str(ident_2) +'.xlsx')
        file_from_3 = ('respuestas/encuesta2/0_'+ str(ident_2) +'.xlsx')
        file_to_3 = ('/respuestas2/0_'+ str(ident_2) +'.xlsx')
        
# Change html pages when press button continue
        list_pages = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
        for i in list_pages:
            if i == button2:
                return render_template('exam_' + button + '.html')


if __name__ == '__main__':
    app.run(host="127.0.0.1", port=5000, debug = True)
